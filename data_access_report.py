import os
import re
import csv
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

def clean_name(s: str) -> str:
    """Trim whitespace and strip any '*' characters."""
    return (s or "").replace("*", "").strip()

def parse_csv_file(path: str) -> list[tuple[str, str]]:
    """
    Exactly replicate what your JS did with Papa.parse(...):
      1. Read all rows with csv.reader
      2. Drop the very first row (assumed to be header)
      3. For each remaining row, grab (last, first, card), trimmed
      4. Filter out rows where both last and first are empty
      5. Group by lowercase key = "last first"
         - If a group has length=1, emit that (last, first).
         - If length>1 and at least one has non‐empty 'card', emit each (last,first) from those that have a card.
         - Otherwise emit the first one.
      6. **Do not sort** the final list—leave it in the order that “uniqueMap” keys were first seen, and within “withCard” in the order they appeared in CSV.
    """
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)

    # 1) drop header
    raw = []
    for row in rows[1:]:
        last = row[0].strip() if len(row) > 0 else ""
        first = row[1].strip() if len(row) > 1 else ""
        card = row[2].strip() if len(row) > 2 else ""
        if last or first:
            raw.append((last, first, card))

    # 2) group by lowercase key
    unique_map: dict[str, list[tuple[str, str, str]]] = {}
    for last, first, card in raw:
        key = f"{last.lower()} {first.lower()}"
        if key not in unique_map:
            unique_map[key] = [(last, first, card)]
        else:
            unique_map[key].append((last, first, card))

    # 3) replicate JS’s logic for duplicates
    result: list[tuple[str, str]] = []
    for key, entries in unique_map.items():
        if len(entries) == 1:
            # Exactly one → take it
            last, first, _ = entries[0]
            result.append((last, first))
        else:
            # More than one → see which have a non‐empty card
            with_card = [e for e in entries if e[2].strip()]
            if with_card:
                # Emit each (last, first) from those that have a card,
                # in the exact order they appeared in CSV.
                for (last, first, _) in with_card:
                    result.append((last, first))
            else:
                # No entries had a card → just emit the first one
                last, first, _ = entries[0]
                result.append((last, first))

    # **Do NOT sort** before returning; JS never sorted csvNames explicitly.
    return result

def main():
    # Hide the Tk root window
    root = tk.Tk()
    root.withdraw()

    # --- 1) Pop‐up #1: pick the main Excel file ---
    excel_path = filedialog.askopenfilename(
        title="Select main Excel file",
        filetypes=[("Excel Workbook", "*.xlsx *.xls")]
    )
    if not excel_path:
        print("No Excel file selected. Exiting.")
        return

    # --- 2) Pop‐up #2: pick ALL door‐CSV files at once ---
    csv_paths = filedialog.askopenfilenames(
        title="Select all door CSV files",
        filetypes=[("CSV files", "*.csv")]
    )
    if not csv_paths:
        print("No CSV files selected. Exiting.")
        return

    # Load the entire workbook
    wb = load_workbook(excel_path)
    sheetnames = set(wb.sheetnames)

    # Build a dict: door_number → list of (last, first) from that CSV
    csv_data: dict[str, list[tuple[str, str]]] = {}
    for path in csv_paths:
        filename = os.path.basename(path)
        m = re.search(r"(\d+)", filename)
        if not m:
            print(f"Skipping '{filename}': no integer found to infer door number.")
            continue
        door_number = m.group(1)
        names = parse_csv_file(path)
        csv_data[door_number] = names

    # For each door_number, find "Door <door_number>" sheet and apply the same merge logic
    for door_number, csv_names in csv_data.items():
        sheet_name = f"Door {door_number}"
        if sheet_name not in sheetnames:
            print(f"No sheet named '{sheet_name}' → skipping.")
            continue

        ws = wb[sheet_name]
        # Extract ALL cell‐values row by row
        all_rows = list(ws.values)  # each row is a tuple of cell‐values

        # Keep rows 0..2 (the first three) exactly as-is (header)
        header_rows = all_rows[:3]
        working_rows = all_rows[3:]  # everything from row 4 onward

        # Build left_names in “sheet order,” exactly how JS did:
        #   leftNames = workingRows.map(row => [ clean(row[3]), clean(row[4]) ]).filter(...)
        left_names: list[tuple[str, str]] = []
        for row in working_rows:
            raw_last = row[3] if len(row) > 3 else ""
            raw_first = row[4] if len(row) > 4 else ""
            last = clean_name(str(raw_last)) if raw_last is not None else ""
            first = clean_name(str(raw_first)) if raw_first is not None else ""
            if last or first:
                left_names.append((last, first))

        # Now run the exact same “one‐pass merge” loop as in JS:
        aligned: list[dict[str, object]] = []
        iL = 0
        iR = 0
        while iL < len(left_names) or iR < len(csv_names):
            l = left_names[iL] if iL < len(left_names) else ("", "")
            r = csv_names[iR]   if iR < len(csv_names)   else ("", "")

            lName = f"{l[0].lower()} {l[1].lower()}".strip()
            rName = f"{r[0].lower()} {r[1].lower()}".strip()

            if iL < len(left_names) and iR < len(csv_names) and lName == rName:
                # exact match
                aligned.append({ "A": l[0], "B": l[1], "D": r[0], "E": r[1], "highlight": None })
                iL += 1
                iR += 1
            elif iR >= len(csv_names) or (iL < len(left_names) and lName < rName):
                # sheet‐entry comes “before” CSV entry → it was REMOVED
                aligned.append({ "A": l[0], "B": l[1], "D": "",   "E": "",   "highlight": "removed" })
                iL += 1
            else:
                # CSV entry comes “before” sheet‐entry → it was ADDED
                aligned.append({ "A": "",    "B": "",    "D": r[0], "E": r[1], "highlight": "added" })
                iR += 1

        # JS then did a pass: if A/B are blank BUT D/E are non‐blank AND highlight is still null,
        # set highlight = "added".  We replicate exactly:
        for obj in aligned:
            aBlank = (not obj["A"]) or (obj["A"].strip() == "")
            bBlank = (not obj["B"]) or (obj["B"].strip() == "")
            dFilled = bool(obj["D"] and obj["D"].strip() != "")
            eFilled = bool(obj["E"] and obj["E"].strip() != "")
            if aBlank and bBlank and dFilled and eFilled and (obj["highlight"] is None):
                obj["highlight"] = "added"

        # Build finalRows exactly as JS’s finalRows = aligned.map(...)
        final_rows: list[list[str]] = []
        for obj in aligned:
            row = [""] * 10
            if obj["highlight"] == "removed":
                # JS: row[0] = A + "**"; row[1] = B + "**"; row[3]=D; row[4]=E
                row[0] = obj["A"] + "**"
                row[1] = obj["B"] + "**"
                row[3] = obj["D"]  # will be ""
                row[4] = obj["E"]  # will be ""
            elif obj["highlight"] == "added":
                # JS: row[0]=A (==""), row[1]=B (==""), row[3]=D + "*", row[4]=E + "*"
                row[0] = obj["A"]
                row[1] = obj["B"]
                row[3] = obj["D"] + "*"  # append exactly one "*"
                row[4] = obj["E"] + "*"
            else:
                # JS’s “matched” branch (highlight === null):
                #   row[0]=A, row[1]=B, row[3]=D, row[4]=E
                row[0] = obj["A"]
                row[1] = obj["B"]
                row[3] = obj["D"]
                row[4] = obj["E"]
            final_rows.append(row)

        # Now wipe out the old sheet completely, then rewrite:
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in r:
                cell.value = None

        # 1) Rewrite header rows 1..3 exactly as they were
        for r_idx, header_row in enumerate(header_rows, start=1):
            for c_idx, cell_val in enumerate(header_row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=cell_val)

        # 2) Starting at row 4, write each of final_rows into columns A→J
        for r_offset, new_row in enumerate(final_rows, start=4):
            for c_offset, val in enumerate(new_row, start=1):
                ws.cell(row=r_offset, column=c_offset, value=val)

        print(f"Processed sheet '{sheet_name}' (door {door_number}).")

    # Finally, decide on output filename exactly as your JS did:
    orig_basename = os.path.basename(excel_path)
    m = re.match(r"^(\d{4})_(\d{2})", orig_basename)
    if m:
        year = int(m.group(1))
        month = int(m.group(2)) + 1
        if month > 12:
            month = 1
            year += 1
        new_name = f"{year:04d}_{month:02d}_Data Center Security List.xlsx"
    else:
        new_name = "Updated_Access_Report.xlsx"

    wb.save(new_name)
    print(f"\n✅ Saved updated report as '{new_name}'.\n")


if __name__ == "__main__":
    main()
