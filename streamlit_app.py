import streamlit as st
import io
import re
import csv
from openpyxl import load_workbook

# --- Helper functions ---
def clean_name(s: str) -> str:
    """Trim whitespace and strip any '*' characters."""
    return (s or "").replace("*", "").strip()


def parse_csv_file(uploaded) -> list[tuple[str, str]]:
    """
    Parse an uploaded CSV (BytesIO) and apply the grouping logic:
      1. Skip header row
      2. Collect (last, first, card)
      3. Remove rows with both last & first empty
      4. Group by lowercase key
      5. For each group:
         - if single → take it
         - if multiple & some have card → emit those with card
         - else → emit first
    """
    text = uploaded.getvalue().decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    raw = []
    for row in rows[1:]:  # drop header
        last = row[0].strip() if len(row) > 0 else ""
        first = row[1].strip() if len(row) > 1 else ""
        card = row[2].strip() if len(row) > 2 else ""
        if last or first:
            raw.append((last, first, card))

    unique_map: dict[str, list[tuple[str, str, str]]] = {}
    for last, first, card in raw:
        key = f"{last.lower()} {first.lower()}"
        unique_map.setdefault(key, []).append((last, first, card))

    result: list[tuple[str, str]] = []
    for entries in unique_map.values():
        if len(entries) == 1:
            result.append((entries[0][0], entries[0][1]))
        else:
            with_card = [e for e in entries if e[2].strip()]
            if with_card:
                for last, first, _ in with_card:
                    result.append((last, first))
            else:
                result.append((entries[0][0], entries[0][1]))
    return result


def process_workbook(excel_bytes: bytes, csv_uploads) -> io.BytesIO:
    """
    Load workbook, align each "Door X" sheet with its corresponding CSV,
    and return a BytesIO of the updated workbook.
    """
    wb = load_workbook(filename=io.BytesIO(excel_bytes))
    sheetnames = set(wb.sheetnames)

    # Map door_number -> parsed CSV names
    csv_data: dict[str, list[tuple[str, str]]] = {}
    for up in csv_uploads:
        m = re.search(r"(\d+)", up.name)
        if not m:
            st.warning(f"Skipping '{up.name}': no door number found.")
            continue
        door = m.group(1)
        names = parse_csv_file(up)
        csv_data[door] = names

    for door, csv_names in csv_data.items():
        sheet = f"Door {door}"
        if sheet not in sheetnames:
            st.warning(f"Sheet '{sheet}' not found, skipping.")
            continue

        ws = wb[sheet]
        all_rows = list(ws.values)
        headers = all_rows[:3]
        data_rows = all_rows[3:]

        # Extract left-side names
        left = []
        for row in data_rows:
            raw_last = row[3] if len(row) > 3 else ""
            raw_first = row[4] if len(row) > 4 else ""
            last = clean_name(str(raw_last)) if raw_last is not None else ""
            first = clean_name(str(raw_first)) if raw_first is not None else ""
            if last or first:
                left.append((last, first))

        # One-pass merge
        aligned = []
        iL = iR = 0
        while iL < len(left) or iR < len(csv_names):
            l = left[iL] if iL < len(left) else ("", "")
            r = csv_names[iR] if iR < len(csv_names) else ("", "")
            lkey = f"{l[0].lower()} {l[1].lower()}".strip()
            rkey = f"{r[0].lower()} {r[1].lower()}".strip()

            if iL < len(left) and iR < len(csv_names) and lkey == rkey:
                aligned.append({"A": l[0], "B": l[1], "D": r[0], "E": r[1], "highlight": None})
                iL += 1
                iR += 1
            elif iR >= len(csv_names) or (iL < len(left) and lkey < rkey):
                aligned.append({"A": l[0], "B": l[1], "D": "", "E": "", "highlight": "removed"})
                iL += 1
            else:
                aligned.append({"A": "", "B": "", "D": r[0], "E": r[1], "highlight": None})
                iR += 1

        # Mark added rows
        for obj in aligned:
            aBlank = not obj["A"].strip()
            bBlank = not obj["B"].strip()
            dFill = bool(obj["D"].strip())
            eFill = bool(obj["E"].strip())
            if aBlank and bBlank and dFill and eFill and obj["highlight"] is None:
                obj["highlight"] = "added"

        # Build final rows
        final = []
        for obj in aligned:
            row = [""] * 10
            if obj["highlight"] == "removed":
                row[0] = obj["A"] + "**"
                row[1] = obj["B"] + "**"
                row[3] = obj["D"]
                row[4] = obj["E"]
            elif obj["highlight"] == "added":
                row[3] = obj["D"] + "*"
                row[4] = obj["E"] + "*"
            else:
                row[0] = obj["A"]
                row[1] = obj["B"]
                row[3] = obj["D"]
                row[4] = obj["E"]
            final.append(row)

        # Clear sheet
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in r:
                cell.value = None

        # Write headers and new rows
        for r_idx, header in enumerate(headers, start=1):
            for c_idx, val in enumerate(header, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        for i, new_row in enumerate(final, start=4):
            for j, val in enumerate(new_row, start=1):
                ws.cell(row=i, column=j, value=val)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --- Streamlit UI ---
st.title("Data Center Security List Generator (Streamlit)")
st.write("Upload your main Excel report and one or more door CSV files. Click 'Process' to generate the updated report.")

excel_up = st.file_uploader("Main Excel file", type=["xlsx", "xls"], help="Choose the Access Report workbook.")
csv_ups = st.file_uploader("Door CSV files", type=["csv"], accept_multiple_files=True, help="Choose door CSV exports.")

if st.button("Process"):
    if not excel_up or not csv_ups:
        st.error("Please upload both the Excel file and at least one CSV.")
    else:
        with st.spinner("Processing..."):
            result = process_workbook(excel_up.read(), csv_ups)
        st.success("Completed processing.")
        st.download_button(
            label="Download Updated Report",
            data=result,
            file_name="Updated_Access_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
